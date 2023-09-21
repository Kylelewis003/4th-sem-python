import openpyxl

def readata(filename , sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    for row in sheet.iter_rows():
        for i in row:
            print(i.value,end = " ")
        

def writedata(filename , sheetname , new_data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    for row_data in new_data:
        sheet.append(row_data)
        
    workbook.save(filename)
    print("Data written to the spreadsheet")

input_file = "input.xlsx"
output_file = "output.xlsx"
sheetname = "Sheet1"

r = readata(input_file,sheetname)

w = [("kyle",20),("neil",23),("Wilma",53)]
writedata(output_file,sheetname,w)
